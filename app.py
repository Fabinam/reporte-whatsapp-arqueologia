import re
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ==========================================================
# CONFIGURACIÓN GENERAL
# ==========================================================

st.set_page_config(
    page_title="Reporte WhatsApp Arqueología",
    page_icon="📋",
    layout="wide",
)

MATERIAL_MAP = {
    "loza": "loza",
    "lozas": "loza",
    "ceramica": "cerámica",
    "ceramicas": "cerámica",
    "vidrio": "vidrio",
    "vidrios": "vidrio",
    "metal": "metal",
    "metales": "metal",
    "osteofauna": "osteofauna",
    "osteofaunas": "osteofauna",
    "plastico": "plástico",
    "plasticos": "plástico",
    "malacologico": "malacológico",
    "malacologicos": "malacológico",
    "miscelaneo": "misceláneo",
    "miscelaneos": "misceláneo",
    "baldosa": "baldosa",
    "baldosas": "baldosa",
}

ORDEN_DESCRIPCION = [
    "vidrio",
    "loza",
    "metal",
    "osteofauna",
    "cerámica",
    "plástico",
    "baldosa",
]

ORDEN_FRECUENCIA = [
    "loza",
    "metal",
    "baldosa",
    "cerámica",
    "vidrio",
    "osteofauna",
    "misceláneo",
]

ORDEN_CRONOLOGIAS = [
    "histórico",
    "prehispánico",
    "subactual",
]

CRONOLOGIAS_OBLIGATORIAS = [
    "subactual",
    "histórico",
    "prehispánico",
]

MATERIAL_SIN_MATERIAL = "sin material"

MATERIALES_EXCLUIDOS_DESCRIPCION = {
    "malacológico",
    "misceláneo",
    MATERIAL_SIN_MATERIAL,
}

COLUMNAS_EDITOR = [
    "fecha",
    "unidad",
    "nivel",
    "estrato",
    "cronologia",
    "material",
    "cantidad",
    "material_nuevo",
    "linea_original",
]

COLUMNAS_COMENTARIOS = [
    "Fecha",
    "Remitente",
    "Unidad",
    "Nivel",
    "Comentario",
]

COLUMNAS_RESUMEN = [
    "Fecha",
    "Unidad",
    "Nivel de inicio",
    "Nivel de término",
    "Total niveles excavados",
    "Nivel de cierre",
    "Estratigrafía",
    "Material cultural",
    "Descripción general",
    "Cronología",
    "Frecuencia históricos",
    "Frecuencia prehispánicos",
]

REPORTE_TIPO = """[dd/mm, hh:mm] Nombre: dd/mm/aaaa

Equipo y comentario breve de la jornada.

Unidad 18-4

Nivel 38, capa C

Subactual: 0

Histórico: 3
-1 osteofauna
-2 lozas

Prehispánico: 0
"""

REPORTE_EJEMPLO = """[10/7, 6:43 p.m.] Nombre: 10/07/2026

Equipo Vale y Clara.
Jornada inicia a las 9:10 hrs.

Unidad 18-4

Nivel 38, capa C

Subactual: 0

Histórico: 3
-1 osteofauna
-2 lozas

Prehispánico: 1
-1 cerámica
"""

FORMATO_REPORTE_TXT = """INSTRUCTIVO PARA EL ENVÍO DE REPORTES ARQUEOLÓGICOS
===================================================

OBJETIVO
-------
Este formato permite que la aplicación reconozca correctamente cada unidad,
nivel, cronología, material y cantidad. Seguir la estructura evita registros
incompletos, alertas innecesarias y correcciones manuales en el Excel.

IMPORTANTE
----------
El reporte se procesa línea por línea. Por esa razón, cada encabezado y cada
material debe escribirse en una línea independiente. No se deben omitir
cronologías: cuando una categoría no tenga materiales, debe informarse con 0.

ORDEN OBLIGATORIO POR CADA NIVEL
--------------------------------
1. Unidad
2. Nivel y capa
3. Subactual
4. Histórico
5. Prehispánico

FORMATO TIPO
------------
[10/7, 6:43 p.m.] Nombre: 10/07/2026

Equipo y comentario breve de la jornada.

Unidad 18-4

Nivel 38, capa C

Subactual: 0

Histórico: 3
-1 osteofauna
-2 lozas

Prehispánico: 1
-1 cerámica

REGLAS DE ENVÍO
---------------
1. Escribir la unidad como:
   Unidad 18
   Unidad 18-4
   Unidad 7a
   Unidad 9-r41

2. Escribir el nivel y la capa como:
   Nivel 38, capa C
   También se acepta: N38, capa C

3. Incluir siempre las tres cronologías en este orden:
   Subactual
   Histórico
   Prehispánico

4. Informar el total junto al encabezado:
   Subactual: 0
   Histórico: 3
   Prehispánico: 1

5. Si no hay materiales en una cronología, escribir 0. No dejar el campo vacío:
   Correcto: Subactual: 0
   Incorrecto: Subactual:
   Incorrecto: omitir Subactual

6. Si el total es mayor que 0, detallar cada material en una línea:
   Histórico: 4
   -2 lozas
   -1 vidrio
   -1 cerámica

7. La suma de los materiales debe coincidir con el total declarado.
   En el ejemplo anterior: 2 + 1 + 1 = 4.

8. Usar idealmente el formato:
   -1 loza
   -3 vidrios
   -1 osteofauna

   También se acepta:
   - Loza: 1
   1 loza

9. Las observaciones de un material pueden escribirse después de una coma:
   -1 osteofauna, fragmento con huella de corte manual

10. Los comentarios generales de jornada deben ir después del encabezado y
    antes de la línea Unidad. Pueden incluir horarios, nombres y actividades.

11. Para más de una unidad, repetir el bloque completo desde “Unidad”.

12. No escribir unidad, nivel, cronología y materiales en una misma línea.

13. No reemplazar los nombres Subactual, Histórico o Prehispánico por nombres
    personales o abreviaturas.

14. Si aparece un material no habitual, escribirlo de todos modos. La
    aplicación lo conservará y lo marcará para revisión.

EJEMPLO CON DOS NIVELES
-----------------------
[10/7, 6:43 p.m.] Nombre: 10/07/2026

Equipo Vale y Clara.
Jornada sin novedades.

Unidad 18-4

Nivel 38, capa C

Subactual: 0

Histórico: 3
-1 osteofauna
-2 lozas

Prehispánico: 1
-1 cerámica

Nivel 39, capa C

Subactual: 0

Histórico: 6
-4 lozas
-2 vidrios

Prehispánico: 0

¿POR QUÉ ES IMPORTANTE RESPETAR ESTE FORMATO?
---------------------------------------------
Cuando una cronología se omite, la aplicación no puede saber si realmente no
hubo materiales o si el dato quedó sin informar. Cuando los totales no
coinciden con el detalle, el Excel puede contener información inconsistente.
Seguir un único formato permite comparar jornadas, reducir errores y mantener
un registro uniforme para todo el equipo.
"""


@dataclass
class Registro:
    fecha: Optional[str]
    unidad: Optional[str]
    nivel: int
    estrato: Optional[str]
    cronologia: str
    material: str
    cantidad: Optional[int]
    material_nuevo: bool = False
    linea_original: str = ""


# ==========================================================
# UTILIDADES DE NORMALIZACIÓN
# ==========================================================

def sin_acentos_minusculas(texto):
    texto = str(texto).strip().lower()
    return "".join(
        caracter
        for caracter in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caracter) != "Mn"
    )


def limpiar_material(texto):
    texto = sin_acentos_minusculas(texto)
    texto = re.sub(r"[^a-zñ ]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_material(texto):
    clave = limpiar_material(texto)

    if not clave:
        return None, False

    material = MATERIAL_MAP.get(clave)

    if material:
        return material, False

    # Los materiales nuevos se conservan, pero no se aceptan frases largas.
    if len(clave.split()) <= 3:
        return clave, True

    return None, False


def normalizar_fecha(dia, mes, anio=None):
    if anio is None:
        anio = "2026"
    else:
        anio = str(anio)
        if len(anio) == 2:
            anio = "20" + anio

    return f"{int(dia):02d}-{int(mes):02d}-{anio}"


def extraer_fecha_linea(linea):
    """Reconoce únicamente fechas que ocupan la línea completa."""
    linea = linea.strip()

    # Con barra se acepta dd/mm o dd/mm/aaaa.
    match = re.fullmatch(
        r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?",
        linea,
    )

    if match:
        return normalizar_fecha(
            match.group(1),
            match.group(2),
            match.group(3),
        )

    # Con guion se exige año para no confundir 18-4 con una fecha.
    match = re.fullmatch(
        r"(\d{1,2})-(\d{1,2})-(\d{2,4})",
        linea,
    )

    if match:
        return normalizar_fecha(
            match.group(1),
            match.group(2),
            match.group(3),
        )

    return None


def extraer_fecha_en_texto(texto):
    """Busca una fecha completa dentro de una alerta como Conteo pendiente."""
    match = re.search(
        r"(?<!\d)(\d{1,2})/(\d{1,2})/(\d{2,4})(?!\d)",
        texto,
    )

    if not match:
        match = re.search(
            r"(?<!\d)(\d{1,2})-(\d{1,2})-(\d{2,4})(?!\d)",
            texto,
        )

    if not match:
        return None

    return normalizar_fecha(
        match.group(1),
        match.group(2),
        match.group(3),
    )


def nivel_a_profundidad_inicio(nivel):
    try:
        nivel = int(nivel)
    except (TypeError, ValueError):
        return "No informado"

    if nivel == 1:
        return "Superficie"

    return f"{nivel} ({(nivel - 1) * 10}-{nivel * 10} cm)"


def nivel_a_profundidad_termino(nivel):
    try:
        nivel = int(nivel)
    except (TypeError, ValueError):
        return "No informado"

    return f"{nivel} ({(nivel - 1) * 10}-{nivel * 10} cm)"


def titulo_material(material):
    if pd.isna(material) or str(material).strip() == "":
        return "No informado"

    return str(material).strip().capitalize()


def nombre_cronologia(cronologia):
    nombres = {
        "histórico": "Histórico",
        "prehispánico": "Prehispánico",
        "subactual": "Subactual",
    }
    return nombres.get(cronologia, str(cronologia).capitalize())


# ==========================================================
# RECONOCIMIENTO ESTRICTO DE LÍNEAS
# ==========================================================

def preparar_lineas(texto):
    texto = texto.replace("\u202f", " ")
    texto = texto.replace("\u2060", "")
    texto = texto.replace("•", "-").replace("·", "-")
    texto = texto.replace("–", "-").replace("—", "-")

    return [
        re.sub(r"\s+", " ", linea.strip())
        for linea in texto.splitlines()
        if linea.strip()
    ]


def reconocer_encabezado_whatsapp(linea):
    match = re.fullmatch(
        r"\[(\d{1,2})/(\d{1,2}),[^\]]+\]\s*([^:]+):\s*(.*)",
        linea,
    )

    if not match:
        return None

    fecha = normalizar_fecha(match.group(1), match.group(2))
    remitente = match.group(3).strip() or "No informado"
    contenido_final = match.group(4).strip()

    # Una fecha escrita después del nombre tiene prioridad por incluir año.
    fecha_final = extraer_fecha_linea(contenido_final)
    if fecha_final:
        fecha = fecha_final
        contenido_final = ""

    return {
        "fecha": fecha,
        "remitente": remitente,
        "contenido_final": contenido_final,
    }


def reconocer_unidad(linea):
    # Formato recomendado: Unidad 18, Unidad 18-4, Unidad 9-r41, Unidad 7a.
    match = re.fullmatch(
        r"Unidad\s+([0-9]+(?:-[A-Za-z0-9]+)?[A-Za-z]?)",
        linea,
        flags=re.I,
    )

    if match:
        return match.group(1).lower()

    # Compatibilidad con reportes antiguos: 9-r41, 8-re o 7a como línea completa.
    # No se acepta 18-4 sin la palabra Unidad para evitar confundirlo con una fecha/rango.
    match = re.fullmatch(
        r"([0-9]+-[A-Za-z][A-Za-z0-9]*|[0-9]+[A-Za-z])",
        linea,
        flags=re.I,
    )

    if match:
        return match.group(1).lower()

    return None


def reconocer_nivel(linea):
    patrones_con_estrato = [
        r"(?:N|Nivel)\s*(\d{1,3})\s*,\s*Capa\s*([A-Za-z]{1,4})",
        r"(?:N|Nivel)\s*(\d{1,3})\s*,\s*([A-Za-z]{1,4})",
        r"(\d{1,3})\s*,\s*([A-Za-z]{1,4})",
    ]

    for patron in patrones_con_estrato:
        match = re.fullmatch(patron, linea, flags=re.I)
        if match:
            return int(match.group(1)), match.group(2).upper(), None

    match = re.fullmatch(
        r"(?:N|Nivel)\s*(\d{1,3})",
        linea,
        flags=re.I,
    )

    if match:
        return int(match.group(1)), None, "Estrato/capa no informado"

    return None


def reconocer_cronologia(linea):
    texto = linea.strip().lstrip("- ").strip()
    texto_sin_acentos = sin_acentos_minusculas(texto)

    patrones = [
        (
            "histórico",
            r"(?:material\s+arqueologico|historico)",
        ),
        (
            "prehispánico",
            r"pre[\s-]*hispanico",
        ),
        (
            "subactual",
            r"(?:material\s+subactual|subactual)",
        ),
    ]

    sufijo = (
        r"\s*(?:"
        r":\s*(\d+)?"
        r"|=\s*(\d+)"
        r"|\(\s*n\s*=\s*(\d+)\s*\)"
        r"|\s+(\d+)"
        r")?\s*"
    )

    for cronologia, patron in patrones:
        match = re.fullmatch(patron + sufijo, texto_sin_acentos)
        if match:
            cantidades = [
                grupo
                for grupo in match.groups()
                if grupo is not None
            ]
            cantidad_declarada = (
                int(cantidades[-1]) if cantidades else None
            )
            return cronologia, cantidad_declarada

    return None


def quitar_vineta(linea):
    return re.sub(r"^[\-]\s*", "", linea.strip()).strip()


def reconocer_materiales(linea):
    """
    Reconoce materiales únicamente cuando la línea completa tiene formato de material.

    Ejemplos:
    -1 loza
    8 lozas
    - Osteofauna: 1, observación
    loza: 2
    loza 2
    """
    original = linea.strip()
    texto = quitar_vineta(original)

    # Varios pares en una misma línea: 1 loza, 3 vidrios.
    segmentos = [segmento.strip() for segmento in re.split(r"[,;]", texto)]
    if len(segmentos) > 1 and all(
        re.fullmatch(
            r"\d+\s+[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+",
            segmento,
        )
        for segmento in segmentos
    ):
        resultados = []
        for segmento in segmentos:
            match = re.fullmatch(
                r"(\d+)\s+([A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+)",
                segmento,
            )
            material, nuevo = normalizar_material(match.group(2))
            if material:
                resultados.append(
                    (
                        material,
                        int(match.group(1)),
                        nuevo,
                        original,
                    )
                )
        return resultados

    # 1 loza / 8 lozas, con observación opcional después de coma o punto y coma.
    match = re.fullmatch(
        r"(\d+)\s+([A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]{1,40}?)"
        r"(?:\s*[,;]\s*(.+))?",
        texto,
    )

    if match:
        material, nuevo = normalizar_material(match.group(2))
        if material:
            return [
                (
                    material,
                    int(match.group(1)),
                    nuevo,
                    original,
                )
            ]

    # Osteofauna: 1, observación / loza: 2.
    match = re.fullmatch(
        r"([A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]{1,40}?)\s*:\s*(\d+)"
        r"(?:\s*[,;]\s*(.+))?",
        texto,
    )

    if match:
        material, nuevo = normalizar_material(match.group(1))
        if material:
            return [
                (
                    material,
                    int(match.group(2)),
                    nuevo,
                    original,
                )
            ]

    # loza 2.
    match = re.fullmatch(
        r"([A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]{1,40}?)\s+(\d+)",
        texto,
    )

    if match:
        material, nuevo = normalizar_material(match.group(1))
        if material:
            return [
                (
                    material,
                    int(match.group(2)),
                    nuevo,
                    original,
                )
            ]

    # material: sin cantidad.
    match = re.fullmatch(
        r"([A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]{1,40}?)\s*:\s*",
        texto,
    )

    if match:
        material, nuevo = normalizar_material(match.group(1))
        if material:
            return [
                (
                    material,
                    None,
                    nuevo,
                    original,
                )
            ]

    return []


def agregar_comentario(
    comentarios,
    fecha,
    remitente,
    unidad,
    nivel,
    texto,
):
    texto = str(texto).strip()

    if not texto:
        return

    registro = {
        "Fecha": fecha or "No informado",
        "Remitente": remitente or "No informado",
        "Unidad": unidad or "No informado",
        "Nivel": nivel if nivel is not None else "No informado",
        "Comentario": texto,
    }

    if registro not in comentarios:
        comentarios.append(registro)


# ==========================================================
# PARSER PRINCIPAL: LÓGICA LÍNEA POR LÍNEA
# ==========================================================

def parsear_todo(texto):
    lineas = preparar_lineas(texto)

    registros = []
    alertas = []
    comentarios = []
    cronologias_declaradas = []

    fecha_actual = None
    remitente_actual = None
    unidad_actual = None
    nivel_actual = None
    estrato_actual = None
    cronologia_actual = None

    for linea in lineas:
        encabezado = reconocer_encabezado_whatsapp(linea)

        if encabezado:
            fecha_actual = encabezado["fecha"]
            remitente_actual = encabezado["remitente"]
            unidad_actual = None
            nivel_actual = None
            estrato_actual = None
            cronologia_actual = None

            if encabezado["contenido_final"]:
                agregar_comentario(
                    comentarios,
                    fecha_actual,
                    remitente_actual,
                    unidad_actual,
                    nivel_actual,
                    encabezado["contenido_final"],
                )
            continue

        fecha_linea = extraer_fecha_linea(linea)

        if fecha_linea:
            fecha_actual = fecha_linea
            continue

        if sin_acentos_minusculas(linea).startswith("conteo pendiente"):
            fecha_pendiente = extraer_fecha_en_texto(linea)
            if fecha_pendiente:
                fecha_actual = fecha_pendiente

            alertas.append(
                "Conteo pendiente: unidad "
                f"{unidad_actual or 'No informado'}, nivel "
                f"{nivel_actual if nivel_actual is not None else 'No informado'}"
            )
            continue

        unidad = reconocer_unidad(linea)

        if unidad:
            unidad_actual = unidad
            nivel_actual = None
            estrato_actual = None
            cronologia_actual = None
            continue

        nivel = reconocer_nivel(linea)

        if nivel:
            nivel_actual, estrato_actual, alerta_nivel = nivel
            cronologia_actual = None

            if unidad_actual is None:
                alertas.append(
                    f"Nivel {nivel_actual} sin unidad informada."
                )

            if alerta_nivel:
                alertas.append(
                    f"{alerta_nivel}: unidad "
                    f"{unidad_actual or 'No informado'}, "
                    f"nivel {nivel_actual}"
                )
            continue

        cronologia = reconocer_cronologia(linea)

        if cronologia:
            cronologia_actual, cantidad_declarada = cronologia

            if unidad_actual is None or nivel_actual is None:
                alertas.append(
                    f"Cronología {nombre_cronologia(cronologia_actual)} "
                    "sin unidad o nivel informado."
                )
                continue

            cronologias_declaradas.append(
                {
                    "fecha": fecha_actual or "No informado",
                    "unidad": unidad_actual,
                    "nivel": nivel_actual,
                    "estrato": estrato_actual,
                    "cronologia": cronologia_actual,
                    "cantidad_declarada": cantidad_declarada,
                }
            )
            continue

        materiales = reconocer_materiales(linea)

        if materiales and cronologia_actual is not None:
            for material, cantidad, material_nuevo, original in materiales:
                if unidad_actual is None:
                    alertas.append(
                        f"Material sin unidad: {original}"
                    )
                    continue

                if nivel_actual is None:
                    alertas.append(
                        f"Material sin nivel: unidad {unidad_actual}, {original}"
                    )
                    continue

                if cantidad is None:
                    alertas.append(
                        f"Cantidad no informada: unidad {unidad_actual}, "
                        f"nivel {nivel_actual}, material {material}"
                    )

                if material_nuevo:
                    alertas.append(
                        f"Material nuevo/no catastrado: '{material}' "
                        f"en unidad {unidad_actual}, nivel {nivel_actual}"
                    )

                registros.append(
                    Registro(
                        fecha=fecha_actual or "No informado",
                        unidad=unidad_actual,
                        nivel=nivel_actual,
                        estrato=estrato_actual,
                        cronologia=cronologia_actual,
                        material=material,
                        cantidad=cantidad,
                        material_nuevo=material_nuevo,
                        linea_original=original,
                    )
                )
            continue

        # Una línea con viñeta dentro de una cronología parece material,
        # pero no pudo interpretarse.
        if cronologia_actual is not None and linea.lstrip().startswith("-"):
            alertas.append(
                "Línea de material no reconocida: unidad "
                f"{unidad_actual or 'No informado'}, nivel "
                f"{nivel_actual if nivel_actual is not None else 'No informado'}, "
                f"{linea}"
            )
            continue

        # Todo lo que no cumple una regla estructural se conserva como comentario.
        agregar_comentario(
            comentarios,
            fecha_actual,
            remitente_actual,
            unidad_actual,
            nivel_actual,
            linea,
        )

    # Agrega una fila auxiliar para cronologías declaradas sin materiales.
    claves_con_material = {
        (
            registro.fecha,
            registro.unidad,
            int(registro.nivel),
            registro.estrato,
            registro.cronologia,
        )
        for registro in registros
        if registro.nivel is not None
    }

    for cronologia in cronologias_declaradas:
        clave = (
            cronologia["fecha"],
            cronologia["unidad"],
            int(cronologia["nivel"]),
            cronologia["estrato"],
            cronologia["cronologia"],
        )

        if clave not in claves_con_material:
            registros.append(
                Registro(
                    fecha=clave[0],
                    unidad=clave[1],
                    nivel=clave[2],
                    estrato=clave[3],
                    cronologia=clave[4],
                    material=MATERIAL_SIN_MATERIAL,
                    cantidad=0,
                    material_nuevo=False,
                    linea_original="Cronología declarada sin materiales",
                )
            )

    return registros, alertas, comentarios, cronologias_declaradas


# ==========================================================
# DATAFRAMES, VALIDACIONES Y RESUMEN
# ==========================================================

def registros_a_df(registros):
    if not registros:
        return pd.DataFrame(columns=COLUMNAS_EDITOR)

    return pd.DataFrame(
        [vars(registro) for registro in registros]
    )[COLUMNAS_EDITOR]


def comentarios_a_df(comentarios):
    if not comentarios:
        return pd.DataFrame(columns=COLUMNAS_COMENTARIOS)

    return pd.DataFrame(
        comentarios,
        columns=COLUMNAS_COMENTARIOS,
    )


def limpiar_df_editado(df):
    if df is None or df.empty:
        return pd.DataFrame(columns=COLUMNAS_EDITOR)

    df = df.copy()

    for columna in COLUMNAS_EDITOR:
        if columna not in df.columns:
            df[columna] = None

    df = df[COLUMNAS_EDITOR]

    df["fecha"] = (
        df["fecha"]
        .fillna("No informado")
        .astype(str)
        .str.strip()
    )

    df["unidad"] = (
        df["unidad"]
        .fillna("No informado")
        .astype(str)
        .str.lower()
        .str.strip()
    )

    df["estrato"] = (
        df["estrato"]
        .fillna("No informado")
        .astype(str)
        .str.upper()
        .str.strip()
    )

    df.loc[
        df["estrato"].isin(["", "NONE", "NAN"]),
        "estrato",
    ] = "No informado"

    df["cronologia"] = (
        df["cronologia"]
        .fillna("No informado")
        .astype(str)
        .str.lower()
        .str.strip()
    )

    df["cronologia"] = df["cronologia"].replace(
        {
            "historico": "histórico",
            "prehispanico": "prehispánico",
            "pre-hispanico": "prehispánico",
            "pre-hispánico": "prehispánico",
            "pre hispanico": "prehispánico",
            "pre hispánico": "prehispánico",
        }
    )

    df["material"] = (
        df["material"]
        .fillna(MATERIAL_SIN_MATERIAL)
        .astype(str)
        .str.lower()
        .str.strip()
    )

    df["nivel"] = pd.to_numeric(
        df["nivel"],
        errors="coerce",
    )

    df["cantidad"] = pd.to_numeric(
        df["cantidad"],
        errors="coerce",
    )

    df["material_nuevo"] = (
        df["material_nuevo"]
        .fillna(False)
        .astype(bool)
    )

    df["linea_original"] = (
        df["linea_original"]
        .fillna("")
        .astype(str)
    )

    df = df.dropna(subset=["nivel"])

    if not df.empty:
        df["nivel"] = df["nivel"].astype(int)

    return df


def nombre_material_descripcion(material):
    if material == "baldosa":
        return "Material constructivo"

    return titulo_material(material)


def nombre_material_frecuencia(material):
    nombres = {
        "loza": "Loza",
        "metal": "Metal",
        "baldosa": "baldosa",
        "cerámica": "Cerámica",
        "vidrio": "Vidrio",
        "osteofauna": "Osteofauna",
        "misceláneo": "Misceláneo",
    }

    return nombres.get(
        str(material),
        titulo_material(material),
    )


def formatear_frecuencia(df_cronologia):
    por_material = defaultdict(int)

    for _, fila in df_cronologia.iterrows():
        por_material[fila["material"]] += int(
            fila["cantidad"]
        )

    orden = ORDEN_FRECUENCIA + sorted(
        material
        for material in por_material
        if material not in ORDEN_FRECUENCIA
    )

    return ", ".join(
        f"{nombre_material_frecuencia(material)} "
        f"(N={por_material[material]})"
        for material in orden
        if material in por_material
    ) or "No informado"


def resumen_por_grupo_df(df):
    df = limpiar_df_editado(df)

    if df.empty:
        return pd.DataFrame(columns=COLUMNAS_RESUMEN)

    resumenes = []

    for (fecha, unidad), grupo in df.groupby(
        ["fecha", "unidad"],
        dropna=False,
    ):
        niveles = sorted(
            grupo["nivel"]
            .dropna()
            .astype(int)
            .unique()
        )

        if not niveles:
            continue

        df_con_material = grupo[
            (grupo["material"] != MATERIAL_SIN_MATERIAL)
            & (grupo["cantidad"].notna())
            & (grupo["cantidad"] > 0)
        ]

        historicos = grupo[
            (grupo["cronologia"] == "histórico")
            & (grupo["material"] != MATERIAL_SIN_MATERIAL)
            & (grupo["cantidad"].notna())
            & (grupo["cantidad"] > 0)
        ]

        prehispanicos = grupo[
            (grupo["cronologia"] == "prehispánico")
            & (grupo["material"] != MATERIAL_SIN_MATERIAL)
            & (grupo["cantidad"].notna())
            & (grupo["cantidad"] > 0)
        ]

        por_estrato = defaultdict(int)

        estratos_declarados = [
            estrato
            for estrato in grupo["estrato"].dropna().unique()
            if str(estrato).strip()
            not in {
                "",
                "No informado",
                "NONE",
                "NAN",
            }
        ]

        for estrato in estratos_declarados:
            por_estrato[estrato] += 0

        # Se mantiene la regla previa: estratigrafía suma materiales históricos.
        for _, fila in historicos.iterrows():
            estrato = (
                fila["estrato"]
                if fila["estrato"] != "No informado"
                else "SIN ESTRATO"
            )
            por_estrato[estrato] += int(fila["cantidad"])

        col_estratigrafia = " ".join(
            f"{estrato} (N={total})"
            for estrato, total in por_estrato.items()
        ) or "No informado"

        materiales_presentes = {
            material
            for material in df_con_material["material"]
            .dropna()
            .unique()
            if material not in MATERIALES_EXCLUIDOS_DESCRIPCION
        }

        orden_descripcion = (
            ORDEN_DESCRIPCION
            + sorted(
                material
                for material in materiales_presentes
                if material not in ORDEN_DESCRIPCION
            )
        )

        col_descripcion = ", ".join(
            nombre_material_descripcion(material)
            for material in orden_descripcion
            if material in materiales_presentes
        ) or "No informado"

        cronologias_presentes = {
            cronologia
            for cronologia in grupo["cronologia"]
            .dropna()
            .unique()
            if cronologia not in {"", "No informado"}
        }

        col_cronologia = " / ".join(
            nombre_cronologia(cronologia)
            for cronologia in ORDEN_CRONOLOGIAS
            if cronologia in cronologias_presentes
        ) or "No informado"

        resumenes.append(
            {
                "Fecha": fecha,
                "Unidad": unidad,
                "Nivel de inicio": nivel_a_profundidad_inicio(
                    min(niveles)
                ),
                "Nivel de término": nivel_a_profundidad_termino(
                    max(niveles)
                ),
                "Total niveles excavados": len(niveles),
                "Nivel de cierre": "",
                "Estratigrafía": col_estratigrafia,
                "Material cultural": (
                    "si" if not df_con_material.empty else "no"
                ),
                "Descripción general": col_descripcion,
                "Cronología": col_cronologia,
                "Frecuencia históricos": formatear_frecuencia(
                    historicos
                ),
                "Frecuencia prehispánicos": formatear_frecuencia(
                    prehispanicos
                ),
            }
        )

    return pd.DataFrame(
        resumenes,
        columns=COLUMNAS_RESUMEN,
    )


def validar_df(
    df,
    alertas_parser=None,
    declaraciones=None,
):
    df = limpiar_df_editado(df)
    alertas = list(alertas_parser or [])
    declaraciones = list(declaraciones or [])

    if df.empty:
        return pd.DataFrame(
            {
                "Alerta": [
                    "No hay datos extraídos ni ingresados manualmente."
                ]
            }
        )

    for indice, fila in df.iterrows():
        numero_fila = indice + 1

        if fila["fecha"] == "No informado":
            alertas.append(
                f"Fila editable {numero_fila}: fecha no informada."
            )

        if fila["unidad"] in {"", "No informado"}:
            alertas.append(
                f"Fila editable {numero_fila}: unidad no informada."
            )

        if fila["estrato"] == "No informado":
            alertas.append(
                f"Fila editable {numero_fila}: estrato/capa no informado."
            )

        if fila["cronologia"] == "No informado":
            alertas.append(
                f"Fila editable {numero_fila}: cronología no informada."
            )

        if (
            fila["material"] != MATERIAL_SIN_MATERIAL
            and pd.isna(fila["cantidad"])
        ):
            alertas.append(
                f"Fila editable {numero_fila}: cantidad no informada "
                f"para material '{fila['material']}'."
            )

        if bool(fila["material_nuevo"]):
            alertas.append(
                f"Fila editable {numero_fila}: material nuevo/no "
                f"catastrado '{fila['material']}'."
            )

    # Cada nivel debe incluir Subactual, Histórico y Prehispánico.
    columnas_nivel = ["fecha", "unidad", "nivel", "estrato"]

    for clave, grupo in df.groupby(
        columnas_nivel,
        dropna=False,
    ):
        _, unidad, nivel, _ = clave
        cronologias_presentes = {
            cronologia
            for cronologia in grupo["cronologia"].dropna().unique()
            if cronologia not in {"", "No informado"}
        }

        for cronologia in CRONOLOGIAS_OBLIGATORIAS:
            if cronologia not in cronologias_presentes:
                nombre = nombre_cronologia(cronologia)
                alertas.append(
                    f"{nombre} no informado: unidad {unidad}, "
                    f"nivel {int(nivel)}. Si no hay materiales, "
                    f"escriba '{nombre}: 0'."
                )

    # Revisa el orden cuando se declararon las tres cronologías.
    declaraciones_por_nivel = defaultdict(list)

    for declaracion in declaraciones:
        nivel_declarado = declaracion.get("nivel")

        if nivel_declarado is None:
            continue

        clave = (
            declaracion.get("fecha", "No informado"),
            declaracion.get("unidad", "No informado"),
            int(nivel_declarado),
            declaracion.get("estrato"),
        )
        cronologia = declaracion.get("cronologia")

        if cronologia not in declaraciones_por_nivel[clave]:
            declaraciones_por_nivel[clave].append(cronologia)

    for clave, orden_declarado in declaraciones_por_nivel.items():
        if all(
            cronologia in orden_declarado
            for cronologia in CRONOLOGIAS_OBLIGATORIAS
        ):
            orden_filtrado = [
                cronologia
                for cronologia in orden_declarado
                if cronologia in CRONOLOGIAS_OBLIGATORIAS
            ]

            if orden_filtrado != CRONOLOGIAS_OBLIGATORIAS:
                _, unidad, nivel, _ = clave
                alertas.append(
                    f"Orden de cronologías no recomendado: unidad "
                    f"{unidad}, nivel {nivel}. Use Subactual, "
                    f"Histórico y Prehispánico, en ese orden."
                )

    # Compara el total informado con la suma de materiales.
    for declaracion in declaraciones:
        nivel_declarado = declaracion.get("nivel")

        if nivel_declarado is None:
            continue

        fecha = declaracion.get("fecha", "No informado")
        unidad = declaracion.get("unidad", "No informado")
        nivel = int(nivel_declarado)
        cronologia = declaracion.get("cronologia")
        cantidad_declarada = declaracion.get("cantidad_declarada")
        nombre = nombre_cronologia(cronologia)

        if cantidad_declarada is None:
            alertas.append(
                f"Total no informado en {nombre}: unidad {unidad}, "
                f"nivel {nivel}. Escriba '{nombre}: 0' o "
                f"'{nombre}: N'."
            )
            continue

        coincidencias = df[
            (df["fecha"] == fecha)
            & (df["unidad"] == unidad)
            & (df["nivel"] == nivel)
            & (df["cronologia"] == cronologia)
            & (df["material"] != MATERIAL_SIN_MATERIAL)
        ]

        suma_materiales = int(
            coincidencias["cantidad"].fillna(0).sum()
        )

        if int(cantidad_declarada) != suma_materiales:
            alertas.append(
                f"Total inconsistente en {nombre}: unidad {unidad}, "
                f"nivel {nivel}. Total declarado "
                f"{int(cantidad_declarada)}, suma de materiales "
                f"{suma_materiales}."
            )

    alertas_unicas = list(dict.fromkeys(alertas))

    if not alertas_unicas:
        return pd.DataFrame({"Estado": ["Sin alertas"]})

    return pd.DataFrame({"Alerta": alertas_unicas})


# ==========================================================
# EXCEL Y RESPALDO
# ==========================================================

def aplicar_formato_excel(writer, nombre_hoja):
    worksheet = writer.book[nombre_hoja]

    relleno_encabezado = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )
    fuente_encabezado = Font(bold=True)
    alineacion_encabezado = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    alineacion_cuerpo = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for celda in worksheet[1]:
        celda.fill = relleno_encabezado
        celda.font = fuente_encabezado
        celda.alignment = alineacion_encabezado

    for fila in worksheet.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = alineacion_cuerpo

    for celdas_columna in worksheet.columns:
        letra_columna = get_column_letter(
            celdas_columna[0].column
        )
        largo_maximo = 0

        for celda in celdas_columna:
            if celda.value is not None:
                largo_maximo = max(
                    largo_maximo,
                    len(str(celda.value)),
                )

        worksheet.column_dimensions[letra_columna].width = min(
            max(largo_maximo + 2, 14),
            70,
        )

    for fila in worksheet.iter_rows():
        worksheet.row_dimensions[fila[0].row].height = (
            35 if fila[0].row == 1 else 45
        )


def crear_excel_en_memoria(
    df_resumen,
    df_detalle,
    df_alertas,
    df_comentarios,
):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_resumen.to_excel(
            writer,
            sheet_name="Para copiar",
            index=False,
        )
        df_detalle.to_excel(
            writer,
            sheet_name="Detalle editable",
            index=False,
        )
        df_alertas.to_excel(
            writer,
            sheet_name="Alertas",
            index=False,
        )
        df_comentarios.to_excel(
            writer,
            sheet_name="Comentarios",
            index=False,
        )

        for nombre_hoja in [
            "Para copiar",
            "Detalle editable",
            "Alertas",
            "Comentarios",
        ]:
            aplicar_formato_excel(writer, nombre_hoja)

    output.seek(0)
    return output.getvalue()


def crear_respaldo_zip(
    texto_original,
    df_resumen,
    df_detalle,
    df_alertas,
    df_comentarios,
):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output = BytesIO()

    excel_bytes = crear_excel_en_memoria(
        df_resumen,
        df_detalle,
        df_alertas,
        df_comentarios,
    )

    with zipfile.ZipFile(
        output,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as archivo_zip:
        archivo_zip.writestr(
            f"{timestamp}_reporte_para_copiar.xlsx",
            excel_bytes,
        )
        archivo_zip.writestr(
            f"{timestamp}_texto_original.txt",
            texto_original,
        )
        archivo_zip.writestr(
            f"{timestamp}_resumen_para_copiar.csv",
            df_resumen.to_csv(index=False).encode("utf-8-sig"),
        )
        archivo_zip.writestr(
            f"{timestamp}_detalle_editable.csv",
            df_detalle.to_csv(index=False).encode("utf-8-sig"),
        )
        archivo_zip.writestr(
            f"{timestamp}_alertas.csv",
            df_alertas.to_csv(index=False).encode("utf-8-sig"),
        )
        archivo_zip.writestr(
            f"{timestamp}_comentarios.csv",
            df_comentarios.to_csv(index=False).encode("utf-8-sig"),
        )
        archivo_zip.writestr(
            "INSTRUCTIVO_FORMATO_REPORTE.txt",
            FORMATO_REPORTE_TXT.encode("utf-8-sig"),
        )

    output.seek(0)
    return output.getvalue(), timestamp


# ==========================================================
# INTERFAZ STREAMLIT
# ==========================================================

def inicializar_estado():
    if "texto_reportes" not in st.session_state:
        st.session_state["texto_reportes"] = ""

    if "resultado_generado" not in st.session_state:
        st.session_state["resultado_generado"] = False

    if "df_extraido" not in st.session_state:
        st.session_state["df_extraido"] = pd.DataFrame(
            columns=COLUMNAS_EDITOR
        )

    if "alertas_parser" not in st.session_state:
        st.session_state["alertas_parser"] = []

    if "comentarios_parser" not in st.session_state:
        st.session_state["comentarios_parser"] = []

    if "declaraciones_parser" not in st.session_state:
        st.session_state["declaraciones_parser"] = []


def limpiar_caja_texto():
    st.session_state["texto_reportes"] = ""


def limpiar_todo():
    st.session_state["texto_reportes"] = ""
    st.session_state["resultado_generado"] = False
    st.session_state["df_extraido"] = pd.DataFrame(
        columns=COLUMNAS_EDITOR
    )
    st.session_state["alertas_parser"] = []
    st.session_state["comentarios_parser"] = []
    st.session_state["declaraciones_parser"] = []


def agregar_reporte_tipo():
    texto_actual = st.session_state.get("texto_reportes", "")

    if texto_actual.strip():
        st.session_state["texto_reportes"] = (
            texto_actual.strip() + "\n\n" + REPORTE_TIPO
        )
    else:
        st.session_state["texto_reportes"] = REPORTE_TIPO


def cargar_reporte_ejemplo():
    texto_actual = st.session_state.get("texto_reportes", "")

    if texto_actual.strip():
        st.session_state["texto_reportes"] = (
            texto_actual.strip() + "\n\n" + REPORTE_EJEMPLO
        )
    else:
        st.session_state["texto_reportes"] = REPORTE_EJEMPLO


def main():
    inicializar_estado()

    st.title("📋 Generador de reportes arqueológicos desde WhatsApp")
    st.caption("Versión 3.1 · validación de cronologías y totales")

    st.write(
        "Procesa reportes de WhatsApp, permite revisar los datos "
        "y genera un Excel descargable."
    )

    with st.expander(
        "ℹ️ Formato obligatorio para el equipo",
        expanded=True,
    ):
        st.markdown(
            """
            Cada nivel debe incluir **Subactual, Histórico y Prehispánico**,
            en ese orden. Cuando no haya materiales, se debe escribir `0`.
            Los totales declarados deben coincidir con la suma del detalle.

            ```text
            [10/7, 6:43 p.m.] Nombre: 10/07/2026

            Comentario breve de la jornada.

            Unidad 18-4

            Nivel 38, capa C

            Subactual: 0

            Histórico: 3
            -1 osteofauna
            -2 lozas

            Prehispánico: 1
            -1 cerámica
            ```

            Los comentarios pueden ir antes de la unidad. La unidad, el nivel,
            cada cronología y cada material deben escribirse en líneas
            independientes.
            """
        )

        st.download_button(
            label="Descargar instructivo de formato TXT",
            data=FORMATO_REPORTE_TXT.encode("utf-8-sig"),
            file_name="FORMATO_REPORTE_ARQUEOLOGICO.txt",
            mime="text/plain",
        )


    st.subheader("1. Ingreso de reportes")

    st.text_area(
        "Pega aquí uno o varios reportes de WhatsApp",
        height=360,
        key="texto_reportes",
        placeholder=REPORTE_TIPO,
    )

    columnas_botones = st.columns([1.15, 1, 1, 1, 1])

    with columnas_botones[0]:
        procesar = st.button(
            "Procesar reportes",
            type="primary",
        )

    with columnas_botones[1]:
        st.button(
            "Limpiar caja",
            on_click=limpiar_caja_texto,
        )

    with columnas_botones[2]:
        st.button(
            "Agregar reporte tipo",
            on_click=agregar_reporte_tipo,
        )

    with columnas_botones[3]:
        st.button(
            "Cargar ejemplo",
            on_click=cargar_reporte_ejemplo,
        )

    with columnas_botones[4]:
        st.button(
            "Limpiar todo",
            on_click=limpiar_todo,
        )

    if procesar:
        texto = st.session_state["texto_reportes"]

        if not texto.strip():
            st.warning("Debes pegar al menos un reporte.")
        else:
            (
                registros,
                alertas,
                comentarios,
                declaraciones,
            ) = parsear_todo(texto)

            st.session_state["df_extraido"] = registros_a_df(
                registros
            )
            st.session_state["alertas_parser"] = alertas
            st.session_state["comentarios_parser"] = comentarios
            st.session_state["declaraciones_parser"] = declaraciones
            st.session_state["resultado_generado"] = True

    if not st.session_state["resultado_generado"]:
        return

    st.subheader("2. Revisar y editar datos extraídos")

    st.info(
        "Puedes corregir campos, agregar filas nuevas o eliminar filas. "
        "Cuando falte información, complétala o déjala como "
        "'No informado'."
    )

    df_base = st.session_state["df_extraido"].copy()

    for columna in COLUMNAS_EDITOR:
        if columna not in df_base.columns:
            df_base[columna] = None

    df_base = df_base[COLUMNAS_EDITOR]

    df_editado = st.data_editor(
        df_base,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "fecha": st.column_config.TextColumn("Fecha"),
            "unidad": st.column_config.TextColumn("Unidad"),
            "nivel": st.column_config.NumberColumn(
                "Nivel",
                step=1,
            ),
            "estrato": st.column_config.TextColumn(
                "Estrato/Capa"
            ),
            "cronologia": st.column_config.SelectboxColumn(
                "Cronología",
                options=[
                    "histórico",
                    "prehispánico",
                    "subactual",
                    "No informado",
                ],
            ),
            "material": st.column_config.TextColumn("Material"),
            "cantidad": st.column_config.NumberColumn(
                "Cantidad",
                step=1,
            ),
            "material_nuevo": st.column_config.CheckboxColumn(
                "Material nuevo"
            ),
            "linea_original": st.column_config.TextColumn(
                "Línea original"
            ),
        },
        key="editor_detalle",
    )

    df_editado_limpio = limpiar_df_editado(df_editado)
    df_resumen = resumen_por_grupo_df(df_editado_limpio)
    df_alertas = validar_df(
        df_editado_limpio,
        st.session_state.get("alertas_parser", []),
        st.session_state.get("declaraciones_parser", []),
    )
    df_comentarios = comentarios_a_df(
        st.session_state.get("comentarios_parser", [])
    )

    st.subheader("3. Alertas y validaciones")

    if "Estado" in df_alertas.columns:
        st.success("Sin alertas.")
    else:
        st.warning(
            "Existen alertas o campos que requieren revisión."
        )

    with st.expander(
        "Ver alertas y validaciones",
        expanded=("Alerta" in df_alertas.columns),
    ):
        st.dataframe(
            df_alertas,
            use_container_width=True,
        )

    materiales_nuevos = sorted(
        set(
            df_editado_limpio.loc[
                df_editado_limpio["material_nuevo"] == True,
                "material",
            ].dropna()
        )
    )

    if materiales_nuevos:
        st.subheader("Materiales no reconocidos detectados")
        st.warning(
            "Se conservarán en el reporte, pero conviene revisar su nombre."
        )
        st.write(
            ", ".join(
                titulo_material(material)
                for material in materiales_nuevos
            )
        )

    st.subheader("4. Comentarios detectados")

    if df_comentarios.empty:
        st.info("No se detectaron comentarios adicionales.")
    else:
        st.dataframe(
            df_comentarios,
            use_container_width=True,
        )

    st.subheader("5. Reporte para copiar")

    if df_resumen.empty:
        st.error(
            "No se pudo generar el resumen. "
            "Revisa o completa la tabla editable."
        )
        return

    st.dataframe(
        df_resumen,
        use_container_width=True,
    )

    excel_bytes = crear_excel_en_memoria(
        df_resumen,
        df_editado_limpio,
        df_alertas,
        df_comentarios,
    )

    columnas_descarga = st.columns([1, 1])

    with columnas_descarga[0]:
        st.download_button(
            label="Descargar Excel",
            data=excel_bytes,
            file_name="reporte_para_copiar.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    with columnas_descarga[1]:
        respaldo_bytes, timestamp = crear_respaldo_zip(
            texto_original=st.session_state["texto_reportes"],
            df_resumen=df_resumen,
            df_detalle=df_editado_limpio,
            df_alertas=df_alertas,
            df_comentarios=df_comentarios,
        )

        st.download_button(
            label="Descargar respaldo completo ZIP",
            data=respaldo_bytes,
            file_name=f"respaldo_reporte_{timestamp}.zip",
            mime="application/zip",
        )


if __name__ == "__main__":
    main()
